"""
Hub de comunicación — canal CORREO (MVP).
Ingesta de novedades del ERP (Excel) al pool, adaptación con IA, maquetación HTML.
Solo admin.

Modelo: una ingesta (Publicacion) es solo un registro de auditoría (cuándo se subió
tal Excel). Los desarrollos que trae viven en un "pool" con su propio
estado_comunicacion (pendiente/comunicado/no_comunicar), independiente de la ingesta
de origen. Una Comunicacion agrupa los desarrollos elegidos del pool — pueden venir de
ingestas distintas — y es lo que realmente se adapta con IA y se envía.
"""
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from typing import Optional
import io
import json
import logging

import re

from app.database import get_db
from app.models.user import User
from app.models.account import Account, Contact, ContactChannel
from app.models.comunicacion import (
    Publicacion, Desarrollo, ComunicacionPrompt, Comunicacion, ComunicacionDesarrollo,
)
from app.schemas.comunicacion import (
    DesarrolloResponse, DesarrolloRelacionado, DesarrolloUpdate,
    PublicacionResponse, PublicacionListResponse, PublicacionDetailResponse,
    IngestaResponse, PoolResponse,
    ComunicacionResponse, ComunicacionUpdate, ComunicacionListItem, ComunicacionListResponse,
    AdaptarCorreoRequest,
    PromptResponse, PromptListResponse, PromptCreate, PromptUpdate, FeedbackItem,
)
from app.utils.auth import get_current_user_from_cookie
from app.utils.audit import generate_id, get_utc_now

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/comunicaciones", tags=["Comunicaciones"])
templates = Jinja2Templates(directory="app/templates")


def _require_admin(current_user: User):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")


# Mapeo flexible de cabeceras del Excel -> campo interno
COLUMN_MAP = {
    "actualización": "titulo_crudo",
    "actualizacion": "titulo_crudo",
    "tipo": "tipo",
    "fecha": "fecha",
    "observaciones": "observaciones",
    "módulo": "modulo",
    "modulo": "modulo",
    "origen": "origen",
    "proyecto": "proyecto",
    "id": "bomp_id_raw",
    "mantenimiento": "mantenimiento_raw",
    "id del desarrollo relacionado": "related_bomp_id_raw",
}

_MANTENIMIENTO_TRUE = {"si", "sí", "yes", "true", "1"}
ESTADOS_COMUNICACION = ("pendiente", "comunicado", "no_comunicar")


def _norm_header(h) -> str:
    return str(h or "").strip().lower()


def _parse_mantenimiento(raw) -> bool:
    if raw is None:
        return False
    return str(raw).strip().lower() in _MANTENIMIENTO_TRUE


def _parse_int(raw):
    if raw is None or isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return int(raw)
    s = str(raw).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _build_desarrollo_responses(db: Session, desarrollos: list, incluir_origen: bool = False) -> list:
    """
    Enriquece una lista de Desarrollo con version_previa/versiones_posteriores,
    buscando en TODO el CRM (no solo en la lista dada): el relacionado puede venir de
    otra ingesta o estar en otra Comunicacion. Si incluir_origen, añade también de qué
    ingesta viene cada desarrollo (para el pool, que es cross-ingesta).
    """
    dev_ids = [d.id for d in desarrollos]
    parent_ids = {d.relacionado_con for d in desarrollos if d.relacionado_con}
    parents_by_id = {}
    if parent_ids:
        for par in db.query(Desarrollo).filter(Desarrollo.id.in_(parent_ids)).all():
            parents_by_id[par.id] = par
    children_by_parent_id = {}
    if dev_ids:
        for child in db.query(Desarrollo).filter(Desarrollo.relacionado_con.in_(dev_ids)).all():
            children_by_parent_id.setdefault(child.relacionado_con, []).append(child)

    pub_by_id = {}
    if incluir_origen and desarrollos:
        pub_ids = {d.publicacion_id for d in desarrollos}
        for p in db.query(Publicacion).filter(Publicacion.id.in_(pub_ids)).all():
            pub_by_id[p.id] = p

    out = []
    for d in desarrollos:
        version_previa = None
        if d.relacionado_con:
            par = parents_by_id.get(d.relacionado_con)
            if par:
                version_previa = DesarrolloRelacionado(
                    id=par.id, publicacion_id=par.publicacion_id,
                    bomp_id=par.bomp_id, titulo_crudo=par.titulo_crudo,
                )
        versiones_posteriores = [
            DesarrolloRelacionado(
                id=c.id, publicacion_id=c.publicacion_id,
                bomp_id=c.bomp_id, titulo_crudo=c.titulo_crudo,
            ) for c in children_by_parent_id.get(d.id, [])
        ]
        pub = pub_by_id.get(d.publicacion_id) if incluir_origen else None
        out.append(DesarrolloResponse(
            id=d.id, publicacion_id=d.publicacion_id, bomp_id=d.bomp_id, titulo_crudo=d.titulo_crudo,
            tipo=d.tipo, fecha=d.fecha, observaciones=d.observaciones, modulo=d.modulo,
            origen=d.origen, proyecto=d.proyecto, norma=d.norma,
            mantenimiento=bool(d.mantenimiento), relacionado_con=d.relacionado_con,
            canales=d.canales, estado_comunicacion=d.estado_comunicacion, orden=d.orden,
            version_previa=version_previa, versiones_posteriores=versiones_posteriores,
            publicacion_version_erp=(pub.version_erp if pub else None),
            publicacion_fecha_ingesta=(pub.fecha_ingesta if pub else None),
        ))
    return out


@router.get("/page", response_class=HTMLResponse)
async def comunicaciones_page(
    request: Request,
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)
    return templates.TemplateResponse("comunicaciones.html", {"request": request})


@router.post("/ingesta-excel", response_model=IngestaResponse, status_code=201)
async def ingesta_excel(
    file: UploadFile = File(...),
    version_erp: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """
    Sube el Excel de novedades del ERP y crea una Publicacion (registro de auditoría de
    la ingesta) + N Desarrollos, que entran al pool con su estado_comunicacion inicial.
    Columnas esperadas: ID, Actualización, Tipo, Fecha, Observaciones, Módulo, Origen,
    Proyecto, Mantenimiento, ID del desarrollo relacionado.
    """
    _require_admin(current_user)

    from openpyxl import load_workbook
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="El archivo está vacío")
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El Excel no tiene filas")

    # Mapear cabecera
    header = rows[0]
    col_idx = {}
    for i, h in enumerate(header):
        key = COLUMN_MAP.get(_norm_header(h))
        if key:
            col_idx[key] = i
    if "titulo_crudo" not in col_idx:
        raise HTTPException(
            status_code=400,
            detail="No se encontró la columna 'Actualización'. Revisa la cabecera del Excel.",
        )

    now = get_utc_now()
    pub = Publicacion(
        id=generate_id(),
        version_erp=(version_erp or "").strip() or None,
        fecha_ingesta=now,
        estado="borrador",
        created_by_user_id=current_user.id,
    )
    db.add(pub)
    db.flush()

    # Desarrollos ya existentes en el CRM (de ingestas anteriores), indexados por
    # bomp_id: sirve tanto para no duplicar al reimportar como para resolver relaciones
    # con desarrollos "padre" que llegaron en un Excel anterior.
    existing_by_bomp_id = {
        d.bomp_id: d for d in db.query(Desarrollo).filter(Desarrollo.bomp_id.isnot(None)).all()
    }
    new_by_bomp_id = {}       # bomp_id -> Desarrollo creado en ESTE import (aun sin flush)
    pending_relations = []    # [(Desarrollo, related_bomp_id), ...] a resolver tras el bucle

    n = 0
    n_omitidos = 0
    for ridx, row in enumerate(rows[1:], start=1):
        def cell(field):
            idx = col_idx.get(field)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return val

        titulo = cell("titulo_crudo")
        if titulo is None or str(titulo).strip() == "":
            continue  # saltar filas vacías

        bomp_id = _parse_int(cell("bomp_id_raw"))
        if bomp_id is not None and (bomp_id in existing_by_bomp_id or bomp_id in new_by_bomp_id):
            n_omitidos += 1
            logger.warning(f"[comunicaciones] fila {ridx}: bomp_id {bomp_id} ya existe en el CRM, se omite (no se duplica)")
            continue

        fecha_val = cell("fecha")
        if fecha_val is not None and hasattr(fecha_val, "strftime"):
            fecha_val = fecha_val.strftime("%Y-%m-%d")

        es_mantenimiento = _parse_mantenimiento(cell("mantenimiento_raw"))

        d = Desarrollo(
            id=generate_id(),
            publicacion_id=pub.id,
            bomp_id=bomp_id,
            titulo_crudo=str(titulo).strip(),
            tipo=(str(cell("tipo")).strip() if cell("tipo") else None),
            fecha=(str(fecha_val).strip() if fecha_val else None),
            observaciones=(str(cell("observaciones")).strip() if cell("observaciones") else None),
            modulo=(str(cell("modulo")).strip() if cell("modulo") else None),
            origen=(str(cell("origen")).strip() if cell("origen") else None),
            proyecto=(str(cell("proyecto")).strip() if cell("proyecto") else None),
            mantenimiento=1 if es_mantenimiento else 0,
            canales="correo",   # por defecto candidato a correo; el socio ajusta
            # Mantenimiento no entra por defecto al pool de comunicables; el socio lo
            # reactiva manualmente (pasa a "pendiente") desde la ficha si hace falta.
            estado_comunicacion="no_comunicar" if es_mantenimiento else "pendiente",
            orden=ridx,
        )
        db.add(d)
        n += 1
        if bomp_id is not None:
            new_by_bomp_id[bomp_id] = d

        related_bomp_id = _parse_int(cell("related_bomp_id_raw"))
        if related_bomp_id is not None:
            pending_relations.append((d, related_bomp_id))

    for d, related_bomp_id in pending_relations:
        parent = new_by_bomp_id.get(related_bomp_id) or existing_by_bomp_id.get(related_bomp_id)
        if not parent:
            logger.warning(
                f"[comunicaciones] BOMP {d.bomp_id}: relacionado con {related_bomp_id} "
                f"pero no existe en el CRM todavía"
            )
            continue
        d.relacionado_con = parent.id
        logger.info(f"[comunicaciones] vinculado BOMP {d.bomp_id} -> BOMP {parent.bomp_id}")

    db.commit()
    logger.info(f"[comunicaciones] ingesta excel pub={pub.id} desarrollos={n} omitidos={n_omitidos}")
    return IngestaResponse(publicacion_id=pub.id, n_desarrollos=n, version_erp=pub.version_erp)


@router.get("", response_model=PublicacionListResponse)
def list_publicaciones(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)
    pubs = db.query(Publicacion).order_by(desc(Publicacion.fecha_ingesta)).all()
    counts = dict(
        db.query(Desarrollo.publicacion_id, func.count(Desarrollo.id))
        .group_by(Desarrollo.publicacion_id).all()
    )
    return PublicacionListResponse(
        publicaciones=[
            PublicacionResponse(
                id=p.id, version_erp=p.version_erp, fecha_ingesta=p.fecha_ingesta,
                estado=p.estado, n_desarrollos=counts.get(p.id, 0),
            ) for p in pubs
        ],
        total=len(pubs),
    )


@router.get("/destinatarios-disponibles")
def destinatarios_disponibles(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """
    Lista de posibles destinatarios sacados de las fichas del CRM (cuentas activas +
    email de empresa, y sus contactos activos con email). Para el selector de
    destinatarios del boletin — evita tener que pegar direcciones a mano.

    IMPORTANTE: registrada ANTES de /{publicacion_id} (misma forma de un solo
    segmento) para que FastAPI no la confunda con un ID de publicacion.
    """
    _require_admin(current_user)

    accounts = db.query(Account).filter(Account.status == "active").order_by(Account.name.asc()).all()
    contacts = db.query(Contact).filter(Contact.status == "active").all()
    contacts_by_account = {}
    for c in contacts:
        contacts_by_account.setdefault(c.account_id, []).append(c)

    contact_ids = [c.id for c in contacts]
    channels = (
        db.query(ContactChannel)
        .filter(ContactChannel.contact_id.in_(contact_ids), ContactChannel.type == "email")
        .all()
        if contact_ids else []
    )
    email_by_contact = {}
    for ch in channels:
        if ch.contact_id not in email_by_contact or ch.is_primary:
            email_by_contact[ch.contact_id] = ch.value

    result = []
    for acc in accounts:
        acc_contacts = []
        for c in contacts_by_account.get(acc.id, []):
            email = email_by_contact.get(c.id)
            if email:
                nombre = f"{c.first_name or ''} {c.last_name or ''}".strip() or "(sin nombre)"
                acc_contacts.append({"contact_id": c.id, "nombre": nombre, "email": email})
        result.append({
            "account_id": acc.id,
            "nombre": acc.name,
            "email_empresa": acc.email,
            "contactos": acc_contacts,
        })
    return {"cuentas": result, "total": len(result)}


@router.get("/pool", response_model=PoolResponse)
def get_pool(
    estado: str = "pendiente",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """
    Pool de desarrollos para comunicar, de CUALQUIER ingesta (no solo de una). Por
    defecto los "pendiente"; pasa ?estado=comunicado o ?estado=no_comunicar para ver
    los otros estados (histórico / descartados).

    IMPORTANTE: registrada ANTES de /{publicacion_id} (mismo motivo que
    destinatarios-disponibles).
    """
    _require_admin(current_user)
    if estado not in ESTADOS_COMUNICACION:
        raise HTTPException(status_code=400, detail=f"estado debe ser uno de {ESTADOS_COMUNICACION}")
    desarrollos = db.query(Desarrollo).filter(Desarrollo.estado_comunicacion == estado).all()
    # BOMP id mas alto primero (aproxima "mas reciente") sin depender del formato de fecha del ERP
    desarrollos.sort(key=lambda d: (d.bomp_id or 0), reverse=True)
    responses = _build_desarrollo_responses(db, desarrollos, incluir_origen=True)
    return PoolResponse(desarrollos=responses, total=len(responses))


@router.get("/comunicaciones", response_model=ComunicacionListResponse)
def list_comunicaciones(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """
    Historial de comunicaciones (borrador/adaptado/publicado).

    IMPORTANTE: registrada ANTES de /{publicacion_id} (mismo motivo que
    destinatarios-disponibles y /pool).
    """
    _require_admin(current_user)
    coms = db.query(Comunicacion).order_by(desc(Comunicacion.created_at)).all()
    counts = dict(
        db.query(ComunicacionDesarrollo.comunicacion_id, func.count(ComunicacionDesarrollo.id))
        .group_by(ComunicacionDesarrollo.comunicacion_id).all()
    )
    items = []
    for c in coms:
        asunto = None
        if c.meta:
            try:
                asunto = json.loads(c.meta).get("asunto")
            except Exception:
                pass
        items.append(ComunicacionListItem(
            id=c.id, canal=c.canal, nombre=c.nombre, estado=c.estado, asunto=asunto,
            n_desarrollos=counts.get(c.id, 0), fecha_publicacion=c.fecha_publicacion,
            created_at=c.created_at,
        ))
    return ComunicacionListResponse(comunicaciones=items, total=len(items))


@router.get("/{publicacion_id}", response_model=PublicacionDetailResponse)
def get_publicacion(
    publicacion_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)
    p = db.query(Publicacion).filter(Publicacion.id == publicacion_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Publicación no encontrada")
    desarrollos = (
        db.query(Desarrollo)
        .filter(Desarrollo.publicacion_id == publicacion_id)
        .order_by(Desarrollo.orden.asc())
        .all()
    )
    n = len(desarrollos)
    return PublicacionDetailResponse(
        publicacion=PublicacionResponse(
            id=p.id, version_erp=p.version_erp, fecha_ingesta=p.fecha_ingesta,
            estado=p.estado, n_desarrollos=n,
        ),
        desarrollos=_build_desarrollo_responses(db, desarrollos),
    )


@router.patch("/desarrollos/{desarrollo_id}", response_model=DesarrolloResponse)
def update_desarrollo(
    desarrollo_id: str,
    payload: DesarrolloUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)
    d = db.query(Desarrollo).filter(Desarrollo.id == desarrollo_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Desarrollo no encontrado")
    if payload.norma is not None: d.norma = payload.norma.strip() or None
    if payload.mantenimiento is not None: d.mantenimiento = 1 if payload.mantenimiento else 0
    if payload.relacionado_con is not None: d.relacionado_con = payload.relacionado_con.strip() or None
    if payload.canales is not None: d.canales = payload.canales.strip() or None
    if payload.estado_comunicacion is not None: d.estado_comunicacion = payload.estado_comunicacion
    if payload.titulo_crudo is not None: d.titulo_crudo = payload.titulo_crudo.strip()
    if payload.observaciones is not None: d.observaciones = payload.observaciones
    if payload.orden is not None: d.orden = payload.orden
    db.commit()
    db.refresh(d)
    return DesarrolloResponse(
        id=d.id, publicacion_id=d.publicacion_id, bomp_id=d.bomp_id, titulo_crudo=d.titulo_crudo,
        tipo=d.tipo, fecha=d.fecha, observaciones=d.observaciones, modulo=d.modulo,
        origen=d.origen, proyecto=d.proyecto, norma=d.norma,
        mantenimiento=bool(d.mantenimiento), relacionado_con=d.relacionado_con,
        canales=d.canales, estado_comunicacion=d.estado_comunicacion, orden=d.orden,
    )


def _desarrollo_to_ai_dict(d: Desarrollo, db: Session) -> dict:
    """
    El id interno de relacionado_con no dice nada a la IA: si el desarrollo tiene
    padre (via bomp_id, ver import), se resuelve aqui a titulo/observaciones + si ese
    padre ya esta "comunicado" (para que la IA fusione ambos en un unico item si van
    en el mismo lote, o escriba el segundo como continuacion si el primero ya salio).
    """
    base = {
        "id": d.id,
        "titulo_crudo": d.titulo_crudo,
        "tipo": d.tipo,
        "modulo": d.modulo,
        "proyecto": d.proyecto,
        "origen": d.origen,
        "observaciones": d.observaciones,
        "mantenimiento": bool(d.mantenimiento),
        "norma": d.norma,
    }
    if d.relacionado_con:
        parent = db.query(Desarrollo).filter(Desarrollo.id == d.relacionado_con).first()
        if parent:
            base["relacionado_con_titulo"] = parent.titulo_crudo
            base["relacionado_con_observaciones"] = parent.observaciones
            base["relacionado_ya_comunicado"] = (parent.estado_comunicacion == "comunicado")
    return base


def _loads(v):
    if not v:
        return None
    try:
        return json.loads(v)
    except Exception:
        return None


def _comunicacion_response(db: Session, c: Comunicacion) -> ComunicacionResponse:
    desarrollos = (
        db.query(Desarrollo)
        .join(ComunicacionDesarrollo, ComunicacionDesarrollo.desarrollo_id == Desarrollo.id)
        .filter(ComunicacionDesarrollo.comunicacion_id == c.id)
        .order_by(ComunicacionDesarrollo.orden.asc())
        .all()
    )
    return ComunicacionResponse(
        id=c.id, canal=c.canal, nombre=c.nombre,
        contenido_generado=_loads(c.contenido_generado),
        contenido_editado=_loads(c.contenido_editado),
        estado=c.estado, meta=_loads(c.meta),
        fecha_publicacion=c.fecha_publicacion, created_at=c.created_at,
        desarrollos=_build_desarrollo_responses(db, desarrollos, incluir_origen=True),
    )


@router.post("/adaptar-correo", response_model=ComunicacionResponse)
def adaptar_correo_endpoint(
    payload: AdaptarCorreoRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """
    Crea una Comunicacion nueva a partir de los desarrollos elegidos del pool (pueden
    ser de ingestas distintas) y llama a la IA para adaptarlos.
    """
    _require_admin(current_user)
    encontrados = db.query(Desarrollo).filter(Desarrollo.id.in_(payload.desarrollo_ids)).all()
    by_id = {d.id: d for d in encontrados}
    faltantes = [i for i in payload.desarrollo_ids if i not in by_id]
    if faltantes:
        raise HTTPException(status_code=404, detail=f"Desarrollos no encontrados: {', '.join(faltantes)}")
    # mantener el orden en que se seleccionaron en el pool
    desarrollos_ordenados = [by_id[i] for i in payload.desarrollo_ids]

    para_correo = [d for d in desarrollos_ordenados if not d.canales or "correo" in (d.canales or "")]
    if not para_correo:
        raise HTTPException(status_code=400, detail="Ninguno de los desarrollos seleccionados está marcado para el canal correo")

    from app.utils.comunicaciones_ai import adaptar_correo, build_system_prompt
    activa = (
        db.query(ComunicacionPrompt)
        .filter(ComunicacionPrompt.canal == "correo", ComunicacionPrompt.activa == 1)
        .first()
    )
    sys_prompt = build_system_prompt(
        prompt_text=(activa.prompt_text if activa else None),
        hero_level=(activa.hero_level if activa else 2),
        calibracion_extra=(activa.calibracion if activa else None),
    )
    try:
        contenido = adaptar_correo([_desarrollo_to_ai_dict(d, db) for d in para_correo], system_prompt=sys_prompt)
    except Exception as e:
        logger.error(f"[comunicaciones] adaptar IA error: {e}")
        raise HTTPException(status_code=502, detail=f"Error en la IA: {e}")

    now = get_utc_now()
    com = Comunicacion(
        id=generate_id(), canal="correo", estado="adaptado",
        contenido_generado=json.dumps(contenido, ensure_ascii=False),
        created_by_user_id=current_user.id, created_at=now, updated_at=now,
    )
    com.contenido_editado = com.contenido_generado
    meta = {
        "asunto": contenido.get("asunto", "Novedades BOMP"),
        "destinatarios_to": [], "destinatarios_cc": [], "destinatarios_bcc": [],
    }
    com.meta = json.dumps(meta, ensure_ascii=False)
    db.add(com)
    db.flush()

    for i, d in enumerate(para_correo):
        db.add(ComunicacionDesarrollo(id=generate_id(), comunicacion_id=com.id, desarrollo_id=d.id, orden=i))

    db.commit()
    db.refresh(com)
    logger.info(f"[comunicaciones] adaptado correo comunicacion={com.id} desarrollos={len(para_correo)}")
    return _comunicacion_response(db, com)


@router.get("/comunicacion/{comunicacion_id}", response_model=ComunicacionResponse)
def get_comunicacion(
    comunicacion_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)
    c = db.query(Comunicacion).filter(Comunicacion.id == comunicacion_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Comunicación no encontrada")
    return _comunicacion_response(db, c)


@router.patch("/comunicacion/{comunicacion_id}", response_model=ComunicacionResponse)
def update_comunicacion(
    comunicacion_id: str,
    payload: ComunicacionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """Guarda la edición del socio (bloques editados, asunto, destinatarios, estado)."""
    _require_admin(current_user)
    c = db.query(Comunicacion).filter(Comunicacion.id == comunicacion_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Comunicación no encontrada")
    if payload.contenido_editado is not None:
        c.contenido_editado = json.dumps(payload.contenido_editado, ensure_ascii=False)
    if payload.meta is not None:
        c.meta = json.dumps(payload.meta, ensure_ascii=False)
    if payload.estado is not None:
        c.estado = payload.estado
    if payload.nombre is not None:
        c.nombre = payload.nombre.strip() or None
    c.updated_at = get_utc_now()
    db.commit()
    db.refresh(c)
    return _comunicacion_response(db, c)


@router.delete("/comunicacion/{comunicacion_id}", status_code=204)
def delete_comunicacion(
    comunicacion_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """Descarta un borrador de comunicación. Sus desarrollos vuelven a estar libres en
    el pool (no se toca su estado_comunicacion, que solo cambia a 'comunicado' al enviar)."""
    _require_admin(current_user)
    c = db.query(Comunicacion).filter(Comunicacion.id == comunicacion_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Comunicación no encontrada")
    db.query(ComunicacionDesarrollo).filter(ComunicacionDesarrollo.comunicacion_id == comunicacion_id).delete()
    db.delete(c)
    db.commit()
    return None


def _marcar_publicada(db: Session, c: Comunicacion):
    """
    Común al envío real por SMTP y a "marcar como enviada" manual (copia/pega o .eml):
    pone la comunicación en publicado y cascada estado_comunicacion='comunicado' a
    todos sus desarrollos — así el pool y el contexto que se le da a la IA sobre
    desarrollos relacionados quedan al día pase lo que pase por dónde se mandó.
    """
    c.estado = "publicado"
    c.fecha_publicacion = get_utc_now()
    c.updated_at = get_utc_now()
    desarrollo_ids = [
        row[0] for row in
        db.query(ComunicacionDesarrollo.desarrollo_id).filter(ComunicacionDesarrollo.comunicacion_id == c.id).all()
    ]
    if desarrollo_ids:
        db.query(Desarrollo).filter(Desarrollo.id.in_(desarrollo_ids)).update(
            {"estado_comunicacion": "comunicado"}, synchronize_session=False
        )


@router.post("/comunicacion/{comunicacion_id}/marcar-enviada", response_model=ComunicacionResponse)
def marcar_enviada(
    comunicacion_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """
    Marca la comunicación como enviada SIN pasar por el envío SMTP real — para cuando
    se ha mandado copiando el HTML con formato a Outlook/Gmail o con el .eml
    descargado. Sin esto el sistema nunca se entera de que ya se comunicó.
    """
    _require_admin(current_user)
    c = db.query(Comunicacion).filter(Comunicacion.id == comunicacion_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Comunicación no encontrada")
    _marcar_publicada(db, c)
    db.commit()
    db.refresh(c)
    logger.info(f"[comunicaciones] marcada como enviada (manual) comunicacion={comunicacion_id} por={current_user.email}")
    return _comunicacion_response(db, c)


def _contenido_y_meta(c: Comunicacion):
    contenido = {}
    if c.contenido_editado:
        contenido = json.loads(c.contenido_editado)
    elif c.contenido_generado:
        contenido = json.loads(c.contenido_generado)
    meta = {}
    if c.meta:
        try:
            meta = json.loads(c.meta)
        except Exception:
            meta = {}
    return contenido, meta


@router.get("/comunicacion/{comunicacion_id}/correo/html", response_class=HTMLResponse)
def correo_html(
    comunicacion_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """Devuelve el HTML email-safe maquetado (para preview y copiar)."""
    _require_admin(current_user)
    c = db.query(Comunicacion).filter(Comunicacion.id == comunicacion_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Comunicación no encontrada")
    contenido, meta = _contenido_y_meta(c)
    firma = meta.get("firma") or (current_user.email_signature or "El equipo de ASIC XXI")
    from app.utils.comunicaciones_ai import build_email_html
    from app.config import get_settings
    cfg = get_settings()
    logo_url = (cfg.public_base_url.rstrip("/") + "/static/img/asicxxi_logo.png") if cfg.public_base_url else ""
    html = build_email_html(
        contenido, firma=firma, saludo=meta.get("saludo", "Hola"),
        logo_url=logo_url,
        cta_web=cfg.comunicaciones_cta_web,
        cta_email=cfg.comunicaciones_cta_email,
        cta_tel=cfg.comunicaciones_cta_tel,
    )
    return HTMLResponse(content=html)


@router.get("/comunicacion/{comunicacion_id}/correo/eml")
def correo_eml(
    comunicacion_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """Descarga un .eml con To/CC/BCC + HTML listo para abrir en Outlook."""
    _require_admin(current_user)
    c = db.query(Comunicacion).filter(Comunicacion.id == comunicacion_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Comunicación no encontrada")
    contenido, meta = _contenido_y_meta(c)
    firma = meta.get("firma") or (current_user.email_signature or "El equipo de ASIC XXI")

    from app.utils.comunicaciones_ai import build_email_html
    from app.config import get_settings
    from email.message import EmailMessage
    cfg = get_settings()
    logo_url = (cfg.public_base_url.rstrip("/") + "/static/img/asicxxi_logo.png") if cfg.public_base_url else ""
    html = build_email_html(
        contenido, firma=firma, saludo=meta.get("saludo", "Hola"),
        logo_url=logo_url,
        cta_web=cfg.comunicaciones_cta_web,
        cta_email=cfg.comunicaciones_cta_email,
        cta_tel=cfg.comunicaciones_cta_tel,
    )

    def _join(lst):
        return ", ".join([x for x in (lst or []) if x])

    msg = EmailMessage()
    # X-Unsent:1 hace que Outlook abra el .eml como mensaje NUEVO editable/enviable
    # (en vez de en modo lectura como un correo ya recibido).
    msg["X-Unsent"] = "1"
    msg["Subject"] = meta.get("asunto", "Novedades BOMP")
    to = _join(meta.get("destinatarios_to"))
    cc = _join(meta.get("destinatarios_cc"))
    bcc = _join(meta.get("destinatarios_bcc") or meta.get("destinatarios_extra"))
    if to:
        msg["To"] = to
    if cc:
        msg["Cc"] = cc
    if bcc:
        msg["Bcc"] = bcc
    msg.set_content("Tu cliente de correo no soporta HTML. Abre el mensaje en un cliente compatible.")
    # base64 evita los artefactos '=' de quoted-printable (asicx=i.es) al partir lineas largas
    msg.add_alternative(html, subtype="html", cte="base64")

    eml_bytes = msg.as_bytes()
    return Response(
        content=eml_bytes,
        media_type="message/rfc822",
        headers={"Content-Disposition": 'attachment; filename="novedades_bomp.eml"'},
    )


_PLACEHOLDER_RE = re.compile(r"⟨[^⟩]*⟩")


def _find_placeholders(contenido: dict) -> list:
    text = json.dumps(contenido, ensure_ascii=False)
    return sorted(set(_PLACEHOLDER_RE.findall(text)))


@router.post("/comunicacion/{comunicacion_id}/correo/enviar")
def enviar_correo(
    comunicacion_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """
    Envio real por SMTP (a diferencia de copiar/pegar o .eml). Usa la misma cuenta
    ya configurada para notificaciones/2FA. Bloquea si quedan placeholders sin
    resolver o si no hay destinatarios.
    """
    _require_admin(current_user)
    c = db.query(Comunicacion).filter(Comunicacion.id == comunicacion_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Comunicación no encontrada")

    contenido, meta = _contenido_y_meta(c)

    placeholders = _find_placeholders(contenido)
    if placeholders:
        raise HTTPException(
            status_code=400,
            detail=f"Hay datos sin confirmar antes de enviar: {', '.join(placeholders)}",
        )

    to = meta.get("destinatarios_to") or []
    cc = meta.get("destinatarios_cc") or []
    bcc = meta.get("destinatarios_bcc") or []
    if not (to or cc or bcc):
        raise HTTPException(status_code=400, detail="No hay destinatarios configurados (pestaña Destinatarios)")

    firma = meta.get("firma") or (current_user.email_signature or "El equipo de ASIC XXI")
    from app.utils.comunicaciones_ai import build_email_html
    from app.config import get_settings
    from app.automations.email_service import send_email_multi
    cfg = get_settings()
    logo_url = (cfg.public_base_url.rstrip("/") + "/static/img/asicxxi_logo.png") if cfg.public_base_url else ""
    html = build_email_html(
        contenido, firma=firma, saludo=meta.get("saludo", "Hola"),
        logo_url=logo_url,
        cta_web=cfg.comunicaciones_cta_web,
        cta_email=cfg.comunicaciones_cta_email,
        cta_tel=cfg.comunicaciones_cta_tel,
    )
    asunto = meta.get("asunto", "Novedades BOMP")

    ok, total = send_email_multi(to=to, subject=asunto, html_body=html, cc=cc, bcc=bcc)
    if not ok:
        raise HTTPException(
            status_code=502,
            detail="No se pudo enviar. Revisa la configuración SMTP (SMTP_USER/SMTP_PASSWORD/EMAIL_ENABLED).",
        )

    _marcar_publicada(db, c)
    db.commit()
    logger.info(f"[comunicaciones] enviado correo comunicacion={comunicacion_id} destinatarios={total} por={current_user.email}")
    return {"message": f"Enviado a {total} destinatarios", "count": total}


@router.delete("/{publicacion_id}", status_code=204)
def delete_publicacion(
    publicacion_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """Borra la ingesta y sus desarrollos. Limpia primero cualquier referencia (bridge
    a comunicaciones, relaciones desde otras ingestas) para no violar las FK."""
    _require_admin(current_user)
    p = db.query(Publicacion).filter(Publicacion.id == publicacion_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="No encontrada")
    dev_ids = [row[0] for row in db.query(Desarrollo.id).filter(Desarrollo.publicacion_id == publicacion_id).all()]
    if dev_ids:
        db.query(ComunicacionDesarrollo).filter(ComunicacionDesarrollo.desarrollo_id.in_(dev_ids)).delete(synchronize_session=False)
        # desarrollos de OTRAS ingestas que apuntaban a estos como padre se quedan sin padre
        db.query(Desarrollo).filter(Desarrollo.relacionado_con.in_(dev_ids)).update(
            {"relacionado_con": None}, synchronize_session=False
        )
    db.query(Desarrollo).filter(Desarrollo.publicacion_id == publicacion_id).delete()
    db.delete(p)
    db.commit()
    return None


# ===========================================================================
# Prompts del adaptador (versiones + hero level + calibración/feedback)
# Router aparte con prefijo más específico, registrado ANTES del router
# principal para que /comunicaciones/prompts no colisione con /{publicacion_id}.
# ===========================================================================

prompts_router = APIRouter(prefix="/comunicaciones/prompts", tags=["Comunicaciones Prompts"])


def _prompt_response(p: ComunicacionPrompt) -> PromptResponse:
    return PromptResponse(
        id=p.id, nombre=p.nombre, canal=p.canal, prompt_text=p.prompt_text,
        hero_level=p.hero_level, calibracion=p.calibracion, activa=bool(p.activa),
    )


@prompts_router.get("", response_model=PromptListResponse)
def list_prompts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)
    from app.utils.comunicaciones_ai import SYSTEM_PROMPT_CORREO
    items = db.query(ComunicacionPrompt).filter(ComunicacionPrompt.canal == "correo").order_by(ComunicacionPrompt.created_at.asc()).all()
    return PromptListResponse(
        prompts=[_prompt_response(p) for p in items],
        default_prompt=SYSTEM_PROMPT_CORREO,
    )


@prompts_router.post("", response_model=PromptResponse, status_code=201)
def create_prompt(
    payload: PromptCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)
    now = get_utc_now()
    p = ComunicacionPrompt(
        id=generate_id(), nombre=payload.nombre.strip(), canal="correo",
        prompt_text=(payload.prompt_text or "").strip() or None,
        hero_level=payload.hero_level, calibracion=(payload.calibracion or "").strip() or None,
        activa=0, created_at=now, updated_at=now,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return _prompt_response(p)


@prompts_router.patch("/{prompt_id}", response_model=PromptResponse)
def update_prompt(
    prompt_id: str,
    payload: PromptUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)
    p = db.query(ComunicacionPrompt).filter(ComunicacionPrompt.id == prompt_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Prompt no encontrado")
    if payload.nombre is not None: p.nombre = payload.nombre.strip()
    if payload.prompt_text is not None: p.prompt_text = payload.prompt_text.strip() or None
    if payload.hero_level is not None: p.hero_level = payload.hero_level
    if payload.calibracion is not None: p.calibracion = payload.calibracion.strip() or None
    p.updated_at = get_utc_now()
    db.commit()
    db.refresh(p)
    return _prompt_response(p)


@prompts_router.post("/{prompt_id}/activar", response_model=PromptResponse)
def activar_prompt(
    prompt_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)
    p = db.query(ComunicacionPrompt).filter(ComunicacionPrompt.id == prompt_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Prompt no encontrado")
    db.query(ComunicacionPrompt).filter(ComunicacionPrompt.canal == "correo").update({"activa": 0})
    p.activa = 1
    p.updated_at = get_utc_now()
    db.commit()
    db.refresh(p)
    return _prompt_response(p)


@prompts_router.delete("/{prompt_id}", status_code=204)
def delete_prompt(
    prompt_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)
    p = db.query(ComunicacionPrompt).filter(ComunicacionPrompt.id == prompt_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="No encontrado")
    db.delete(p)
    db.commit()
    return None


@prompts_router.post("/feedback", status_code=200)
def feedback_calibracion(
    payload: FeedbackItem,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """
    Añade un ejemplo de calibración (bien/meh/mal) al prompt ACTIVO desde la
    pantalla de curación. Si no hay prompt activo, crea uno por defecto editable.
    """
    _require_admin(current_user)
    activa = (
        db.query(ComunicacionPrompt)
        .filter(ComunicacionPrompt.canal == "correo", ComunicacionPrompt.activa == 1)
        .first()
    )
    if not activa:
        now = get_utc_now()
        activa = ComunicacionPrompt(
            id=generate_id(), nombre="Prompt activo (auto)", canal="correo",
            prompt_text=None, hero_level=2, calibracion=None, activa=1,
            created_at=now, updated_at=now,
        )
        db.add(activa)
        db.flush()

    marca = {"bien": "BIEN ✅", "meh": "MEH 😐", "mal": "MAL ❌"}.get(payload.veredicto, payload.veredicto)
    linea = f'- {marca} titulo: "{(payload.titulo or "").strip()}". cuerpo: "{(payload.cuerpo or "").strip()}".'
    if payload.nota:
        linea += f' Nota del revisor: {payload.nota.strip()}'
    activa.calibracion = ((activa.calibracion or "") + "\n" + linea).strip()
    activa.updated_at = get_utc_now()
    db.commit()
    return {"ok": True, "prompt_id": activa.id}
