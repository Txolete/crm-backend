"""
Importador de clientes (cuentas + contactos) desde el export de BOMP.
Formato esperado: Excel con 2 hojas.

Hoja "Clientes": Nombre, CIF/NIF, Email empresa, Telefono, Web, Direccion, Provincia, Notas
Hoja "Contactos": CIF/NIF cliente, Nombre cliente, Nombre, Rol, Email, Telefono, Acceso, Activo

Reutilizable: pensado para volver a correr cada vez que IT mande una extraccion nueva.
Dedupe por CIF/NIF (cuentas) y por email/nombre dentro de la cuenta (contactos).
Nota: CIF/email/telefono/nombre estan cifrados en columna -> comparacion en Python, no SQL.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
import io
import logging

from app.database import get_db
from app.models.user import User
from app.models.account import Account, Contact, ContactChannel
from app.models.config import CfgRegion, CfgContactRole
from app.utils.auth import get_current_user_from_cookie
from app.utils.audit import generate_id, get_utc_now
from app.utils.import_parsers import validate_email, validate_phone, safe_str

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/import-clients", tags=["Import Clients"])
templates = Jinja2Templates(directory="app/templates")


def _require_admin(current_user: User):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")


def _norm_cif(v) -> str:
    return safe_str(v).strip().upper().replace(" ", "") if v else ""


def _split_name(full_name: str):
    parts = (full_name or "").strip().split(None, 1)
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


class RowResult(BaseModel):
    row: int
    status: str  # created | updated | skipped | failed
    entity: str  # cliente | contacto
    name: Optional[str] = None
    detail: Optional[str] = None


class ImportClientsResponse(BaseModel):
    dry_run: bool
    clientes_creados: int
    clientes_actualizados: int
    clientes_omitidos: int
    contactos_creados: int
    contactos_actualizados: int
    contactos_omitidos: int
    detalle: List[RowResult]


@router.get("/page", response_class=HTMLResponse)
async def import_clients_page(request: Request, current_user: User = Depends(get_current_user_from_cookie)):
    _require_admin(current_user)
    return templates.TemplateResponse("import_clients.html", {"request": request})


@router.get("/export")
async def export_clients(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    """
    Extracción de clientes + contactos en el mismo formato (2 hojas) que espera el
    importador, para poder depurar en Excel y volver a subir directamente.
    """
    _require_admin(current_user)

    from openpyxl import Workbook

    regions_by_id = {r.id: r.name for r in db.query(CfgRegion).all()}
    roles_by_id = {r.id: r.name for r in db.query(CfgContactRole).all()}

    accounts = db.query(Account).order_by(Account.status, Account.name).all()
    accounts_by_id = {a.id: a for a in accounts}

    contacts = db.query(Contact).order_by(Contact.account_id).all()
    channels = db.query(ContactChannel).all()
    channels_by_contact = {}
    for ch in channels:
        channels_by_contact.setdefault(ch.contact_id, []).append(ch)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Clientes"
    ws1.append(["Nombre", "CIF/NIF", "Email empresa", "Teléfono", "Web", "Dirección", "Provincia", "Notas", "Estado"])
    for acc in accounts:
        provincia = regions_by_id.get(acc.region_id) or acc.region_other_text or ""
        ws1.append([
            acc.name or "",
            acc.tax_id or "",
            acc.email or "",
            acc.phone or "",
            acc.website or "",
            acc.address or "",
            provincia,
            acc.notes or "",
            "Activo" if acc.status == "active" else "Archivado",
        ])

    ws2 = wb.create_sheet("Contactos")
    ws2.append(["CIF/NIF cliente", "Nombre cliente", "Nombre", "Rol", "Email", "Teléfono", "Acceso", "Activo"])
    for c in contacts:
        acc = accounts_by_id.get(c.account_id)
        if not acc:
            continue
        full_name = f"{c.first_name or ''} {c.last_name or ''}".strip()
        rol = roles_by_id.get(c.contact_role_id) or c.contact_role_other_text or ""
        chs = channels_by_contact.get(c.id, [])
        email = next((ch.value for ch in chs if ch.type == "email" and ch.is_primary), "") \
            or next((ch.value for ch in chs if ch.type == "email"), "")
        phone = next((ch.value for ch in chs if ch.type == "phone"), "")
        ws2.append([
            acc.tax_id or "",
            acc.name or "",
            full_name,
            rol,
            email,
            phone,
            "",
            "Sí" if c.status == "active" else "No",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"extraccion_clientes_crm_{get_utc_now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    logger.info(f"[import-clients] export por {current_user.email}: {len(accounts)} cuentas, {len(contacts)} contactos")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("", response_model=ImportClientsResponse)
async def import_clients(
    file: UploadFile = File(...),
    dry_run: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie),
):
    _require_admin(current_user)

    from openpyxl import load_workbook
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="El archivo está vacío")
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")

    if "Clientes" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="No se encontró la hoja 'Clientes'")

    detalle: List[RowResult] = []
    clientes_creados = clientes_actualizados = clientes_omitidos = 0
    contactos_creados = contactos_actualizados = contactos_omitidos = 0

    # --- Cache de regiones (Provincia) ---
    regions = db.query(CfgRegion).all()
    regions_by_name = {r.name.strip().lower(): r for r in regions}

    # --- Cache de roles de contacto ---
    roles = db.query(CfgContactRole).all()
    roles_by_name = {r.name.strip().lower(): r for r in roles}

    # --- Cargar cuentas existentes en memoria (CIF esta cifrado -> comparar en Python) ---
    existing_accounts = db.query(Account).all()
    accounts_by_cif = {}
    for acc in existing_accounts:
        cif = _norm_cif(acc.tax_id)
        if cif:
            accounts_by_cif[cif] = acc

    # ============================================================
    # HOJA CLIENTES
    # ============================================================
    ws = wb["Clientes"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h or "").strip().lower() for h in rows[0]] if rows else []

    def col(row, name, default=None):
        try:
            idx = header.index(name)
        except ValueError:
            return default
        return row[idx] if idx < len(row) else default

    for ridx, row in enumerate(rows[1:], start=2):
        name = safe_str(col(row, "nombre"))
        if not name:
            continue
        cif = _norm_cif(col(row, "cif/nif"))
        email = safe_str(col(row, "email empresa"))
        phone = safe_str(col(row, "teléfono") or col(row, "telefono"))
        website = safe_str(col(row, "web"))
        address = safe_str(col(row, "dirección") or col(row, "direccion"))
        province = safe_str(col(row, "provincia"))
        notes = safe_str(col(row, "notas"))

        existing = accounts_by_cif.get(cif) if cif else None

        if existing:
            existing.name = name
            if email:
                is_valid, _ = validate_email(email)
                existing.email = email if is_valid else existing.email
            if phone:
                existing.phone = phone
            if website:
                existing.website = website
            if address:
                existing.address = address
            if notes:
                existing.notes = notes
            if province:
                region = regions_by_name.get(province.strip().lower())
                if region:
                    existing.region_id = region.id
                else:
                    existing.region_other_text = province
            existing.updated_at = get_utc_now()
            if not dry_run:
                db.flush()
            clientes_actualizados += 1
            detalle.append(RowResult(row=ridx, status="updated", entity="cliente", name=name))
            accounts_by_cif[cif] = existing
        else:
            acc = Account(
                id=generate_id(),
                name=name,
                tax_id=cif or None,
                status="active",
                created_at=get_utc_now(),
                updated_at=get_utc_now(),
            )
            if email:
                is_valid, _ = validate_email(email)
                if is_valid:
                    acc.email = email
            if phone:
                acc.phone = phone
            if website:
                acc.website = website
            if address:
                acc.address = address
            if notes:
                acc.notes = notes
            if province:
                region = regions_by_name.get(province.strip().lower())
                if region:
                    acc.region_id = region.id
                else:
                    acc.region_other_text = province

            if not dry_run:
                db.add(acc)
                db.flush()
            clientes_creados += 1
            detalle.append(RowResult(row=ridx, status="created", entity="cliente", name=name))
            if cif:
                accounts_by_cif[cif] = acc

    # ============================================================
    # HOJA CONTACTOS
    # ============================================================
    if "Contactos" in wb.sheetnames:
        ws2 = wb["Contactos"]
        rows2 = list(ws2.iter_rows(values_only=True))
        header2 = [str(h or "").strip().lower() for h in rows2[0]] if rows2 else []

        def col2(row, name, default=None):
            try:
                idx = header2.index(name)
            except ValueError:
                return default
            return row[idx] if idx < len(row) else default

        # Cache de contactos ya cargados por cuenta (para dedupe, evita N+1)
        contacts_cache = {}

        for ridx, row in enumerate(rows2[1:], start=2):
            cif = _norm_cif(col2(row, "cif/nif cliente"))
            full_name = safe_str(col2(row, "nombre"))
            role_name = safe_str(col2(row, "rol"))
            email = safe_str(col2(row, "email"))
            phone = safe_str(col2(row, "teléfono") or col2(row, "telefono"))
            activo = safe_str(col2(row, "activo"))

            if not full_name:
                contactos_omitidos += 1
                detalle.append(RowResult(row=ridx, status="skipped", entity="contacto", detail="Sin nombre"))
                continue

            account = accounts_by_cif.get(cif) if cif else None
            if not account:
                contactos_omitidos += 1
                detalle.append(RowResult(
                    row=ridx, status="skipped", entity="contacto", name=full_name,
                    detail=f"Cliente con CIF '{cif}' no encontrado (¿falta en la hoja Clientes?)"
                ))
                continue

            if account.id not in contacts_cache:
                contacts_cache[account.id] = db.query(Contact).filter(Contact.account_id == account.id).all()
            acc_contacts = contacts_cache[account.id]

            first_name, last_name = _split_name(full_name)
            existing_contact = next(
                (c for c in acc_contacts
                 if (c.first_name or "").strip().lower() == first_name.strip().lower()
                 and (c.last_name or "").strip().lower() == last_name.strip().lower()),
                None,
            )

            status_val = "active" if activo.strip().lower() not in ("no", "false", "0") else "archived"

            if existing_contact:
                existing_contact.status = status_val
                if role_name:
                    role = roles_by_name.get(role_name.strip().lower())
                    if role:
                        existing_contact.contact_role_id = role.id
                    else:
                        existing_contact.contact_role_other_text = role_name
                existing_contact.updated_at = get_utc_now()
                contactos_actualizados += 1
                detalle.append(RowResult(row=ridx, status="updated", entity="contacto", name=full_name))
                if not dry_run:
                    db.flush()
                continue

            contact = Contact(
                id=generate_id(),
                account_id=account.id,
                first_name=first_name,
                last_name=last_name,
                status=status_val,
                created_at=get_utc_now(),
                updated_at=get_utc_now(),
            )
            if role_name:
                role = roles_by_name.get(role_name.strip().lower())
                if role:
                    contact.contact_role_id = role.id
                else:
                    contact.contact_role_other_text = role_name

            if not dry_run:
                db.add(contact)
                db.flush()

            if email:
                is_valid, _ = validate_email(email)
                if is_valid:
                    ch = ContactChannel(
                        id=generate_id(), contact_id=contact.id, type="email",
                        value=email, is_primary=1, created_at=get_utc_now(),
                    )
                    if not dry_run:
                        db.add(ch)
            if phone:
                is_valid, _ = validate_phone(str(phone))
                if is_valid:
                    ch = ContactChannel(
                        id=generate_id(), contact_id=contact.id, type="phone",
                        value=str(phone), is_primary=0, created_at=get_utc_now(),
                    )
                    if not dry_run:
                        db.add(ch)

            if not dry_run:
                db.flush()
            contacts_cache[account.id].append(contact)
            contactos_creados += 1
            detalle.append(RowResult(row=ridx, status="created", entity="contacto", name=full_name))

    if not dry_run:
        db.commit()
        logger.info(
            f"[import-clients] commit por {current_user.email}: "
            f"clientes +{clientes_creados}/~{clientes_actualizados}, "
            f"contactos +{contactos_creados}/~{contactos_actualizados}"
        )
    else:
        db.rollback()  # dry-run: descartar cualquier flush intermedio

    return ImportClientsResponse(
        dry_run=dry_run,
        clientes_creados=clientes_creados,
        clientes_actualizados=clientes_actualizados,
        clientes_omitidos=clientes_omitidos,
        contactos_creados=contactos_creados,
        contactos_actualizados=contactos_actualizados,
        contactos_omitidos=contactos_omitidos,
        detalle=detalle,
    )
