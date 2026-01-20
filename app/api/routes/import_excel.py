"""
Excel Import endpoints
PASO 6 - Robust Excel importer with validation and warnings
"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from sqlalchemy.orm import Session
from datetime import datetime
from pathlib import Path
import json
import shutil
from openpyxl import load_workbook

from app.database import get_db
from app.models.user import User
from app.models.account import Account, Contact, ContactChannel
from app.models.opportunity import Opportunity, Task, Activity
from app.models.config import (
    CfgStage, CfgLeadSource, CfgRegion, CfgCustomerType,
    CfgContactRole, CfgTaskTemplate
)
from app.schemas.import_excel import ImportResponse, ImportRowResult, ImportReportResponse
from app.utils.auth import get_current_user_from_cookie, require_role
from app.utils.audit import create_audit_log, generate_id, get_iso_timestamp
from app.utils.import_parsers import (
    parse_date, parse_month, parse_currency, validate_email,
    validate_phone, safe_str, safe_float
)
import logging

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/import", tags=["Import"])

# Upload directory
UPLOAD_DIR = Path("/data/uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def process_excel_row(
    row_data: dict,
    row_num: int,
    db: Session,
    current_user: User,
    dry_run: bool,
    config_cache: dict
) -> ImportRowResult:
    """
    Process a single Excel row
    
    Returns ImportRowResult with status, warnings, errors
    """
    warnings = []
    errors = []
    account_name = safe_str(row_data.get('account_name'))
    opportunity_id = None
    
    # === CRITICAL VALIDATIONS ===
    if not account_name:
        errors.append("Account name is required")
        return ImportRowResult(
            row=row_num,
            status="failed",
            errors=errors,
            warnings=warnings
        )
    
    expected_value_str = row_data.get('expected_value_eur')
    expected_value, err = parse_currency(expected_value_str)
    if expected_value is None or expected_value < 0:
        errors.append(f"Invalid expected_value_eur: {expected_value_str}")
        return ImportRowResult(
            row=row_num,
            status="failed",
            errors=errors,
            warnings=warnings,
            account_name=account_name
        )
    
    stage_key = safe_str(row_data.get('stage'))
    if not stage_key:
        errors.append("Stage is required")
        return ImportRowResult(
            row=row_num,
            status="failed",
            errors=errors,
            warnings=warnings,
            account_name=account_name
        )
    
    # Find stage
    stage = config_cache['stages_by_key'].get(stage_key.lower())
    if not stage:
        errors.append(f"Stage '{stage_key}' not found in configuration")
        return ImportRowResult(
            row=row_num,
            status="failed",
            errors=errors,
            warnings=warnings,
            account_name=account_name
        )
    
    # If there are critical errors, stop here
    if errors:
        return ImportRowResult(
            row=row_num,
            status="failed",
            errors=errors,
            warnings=warnings,
            account_name=account_name
        )
    
    # === PROCESS ACCOUNT (upsert by name) ===
    account = db.query(Account).filter(
        Account.name.ilike(account_name)
    ).first()
    
    if not account:
        # Create new account
        account = Account(
            id=generate_id(),
            name=account_name,
            status='active',
            created_at=get_iso_timestamp(),
            updated_at=get_iso_timestamp()
        )
        if not dry_run:
            db.add(account)
    
    # Update account fields
    region_name = safe_str(row_data.get('region'))
    if region_name:
        region = config_cache['regions_by_name'].get(region_name.lower())
        if region:
            account.region_id = region.id
        else:
            # HOTFIX 6.1: Use region_other if available
            region_other = safe_str(row_data.get('region_other'))
            account.region_other_text = region_other if region_other else region_name
            warnings.append(f"Region '{region_name}' not found, saved to other_text")
    
    customer_type_name = safe_str(row_data.get('customer_type'))
    if customer_type_name:
        customer_type = config_cache['customer_types_by_name'].get(customer_type_name.lower())
        if customer_type:
            account.customer_type_id = customer_type.id
        else:
            # HOTFIX 6.1: Use customer_type_other if available
            customer_type_other = safe_str(row_data.get('customer_type_other'))
            account.customer_type_other_text = customer_type_other if customer_type_other else customer_type_name
            warnings.append(f"Customer type '{customer_type_name}' not found, saved to other_text")
    
    lead_source_name = safe_str(row_data.get('lead_source'))
    if lead_source_name:
        lead_source = config_cache['lead_sources_by_name'].get(lead_source_name.lower())
        if lead_source:
            account.lead_source_id = lead_source.id
        else:
            warnings.append(f"Lead source '{lead_source_name}' not found, no cfg created")
    
    lead_source_detail = safe_str(row_data.get('lead_source_detail'))
    if lead_source_detail:
        account.lead_source_detail = lead_source_detail
    
    owner_email = safe_str(row_data.get('owner_email'))
    if owner_email:
        owner = config_cache['users_by_email'].get(owner_email.lower())
        if owner:
            account.owner_user_id = owner.id
        else:
            warnings.append(f"Owner '{owner_email}' not found")
    
    account.updated_at = get_iso_timestamp()
    
    if not dry_run:
        db.flush()  # CRITICAL: Flush account to DB so FK constraints work
    
    # === PROCESS CONTACT (optional) ===
    contact_first = safe_str(row_data.get('contact_first_name'))
    contact_last = safe_str(row_data.get('contact_last_name'))
    
    if contact_first or contact_last:
        # Check if contact already exists
        existing_contact = db.query(Contact).filter(
            Contact.account_id == account.id,
            Contact.first_name == contact_first,
            Contact.last_name == contact_last
        ).first()
        
        if not existing_contact:
            contact = Contact(
                id=generate_id(),
                account_id=account.id,
                first_name=contact_first,
                last_name=contact_last,
                status='active',
                created_at=get_iso_timestamp(),
                updated_at=get_iso_timestamp()
            )
            
            # Contact role
            contact_role_name = safe_str(row_data.get('contact_role'))
            if contact_role_name:
                contact_role = config_cache['contact_roles_by_name'].get(contact_role_name.lower())
                if contact_role:
                    contact.contact_role_id = contact_role.id
                else:
                    contact.contact_role_other_text = contact_role_name
                    warnings.append(f"Contact role '{contact_role_name}' not found, saved to other_text")
            
            if not dry_run:
                db.add(contact)
            
            # Contact channels (email, phone)
            contact_email = safe_str(row_data.get('contact_email'))
            if contact_email:
                is_valid, err = validate_email(contact_email)
                if is_valid:
                    channel = ContactChannel(
                        id=generate_id(),
                        contact_id=contact.id,
                        type='email',
                        value=contact_email,
                        is_primary=1,
                        created_at=get_iso_timestamp()
                    )
                    if not dry_run:
                        db.add(channel)
                else:
                    warnings.append(f"Invalid email: {contact_email}")
            
            contact_phone = safe_str(row_data.get('contact_phone'))
            if contact_phone:
                is_valid, err = validate_phone(contact_phone)
                if is_valid:
                    channel = ContactChannel(
                        id=generate_id(),
                        contact_id=contact.id,
                        type='phone',
                        value=contact_phone,
                        is_primary=0,
                        created_at=get_iso_timestamp()
                    )
                    if not dry_run:
                        db.add(channel)
                else:
                    warnings.append(f"Invalid phone: {contact_phone}")
            
            if not dry_run:
                db.flush()  # CRITICAL: Flush contact and channels to DB
        else:
            warnings.append(f"Contact {contact_first} {contact_last} already exists, skipped")
    
    # === PROCESS OPPORTUNITY ===
    opportunity = Opportunity(
        id=generate_id(),
        account_id=account.id,
        name=safe_str(row_data.get('opportunity_name')),
        stage_id=stage.id,
        expected_value_eur=expected_value,
        close_outcome='open',
        status='active',
        created_at=get_iso_timestamp(),
        updated_at=get_iso_timestamp()
    )
    
    # HOTFIX 6.1: weighted_value_override_eur (optional)
    weighted_value_str = row_data.get('weighted_value_eur')
    if weighted_value_str:
        weighted_value, err = parse_currency(weighted_value_str)
        if weighted_value is not None:
            opportunity.weighted_value_override_eur = weighted_value
        else:
            warnings.append(f"Invalid weighted_value_eur: {err}")
    
    # Forecast close month
    forecast_month_str = row_data.get('forecast_close_month')
    if forecast_month_str:
        forecast_month, err = parse_month(forecast_month_str)
        if forecast_month:
            opportunity.forecast_close_month = forecast_month
        else:
            warnings.append(f"Invalid forecast_close_month: {err}")
    
    # Close outcome
    close_outcome = safe_str(row_data.get('close_outcome'))
    if close_outcome and close_outcome.lower() in ['open', 'won', 'lost']:
        opportunity.close_outcome = close_outcome.lower()
        
        # If won/lost, need close_date
        close_date_str = row_data.get('close_date')
        if close_date_str:
            close_date, err = parse_date(close_date_str)
            if close_date:
                opportunity.close_date = close_date
            else:
                warnings.append(f"Invalid close_date: {err}")
        
        # If won, capture won_value
        if opportunity.close_outcome == 'won':
            won_value_str = row_data.get('won_value_eur')
            if won_value_str:
                won_value, err = parse_currency(won_value_str)
                if won_value:
                    opportunity.won_value_eur = won_value
        
        # If lost, capture lost_reason
        if opportunity.close_outcome == 'lost':
            lost_reason = safe_str(row_data.get('lost_reason'))
            if lost_reason:
                opportunity.lost_reason = lost_reason
            else:
                warnings.append("Lost opportunity should have lost_reason")
    
    # Owner
    if owner_email and owner_email in config_cache['users_by_email']:
        opportunity.owner_user_id = config_cache['users_by_email'][owner_email.lower()].id
    
    if not dry_run:
        db.add(opportunity)
        db.flush()  # CRITICAL: Flush to DB so FK constraints work for tasks/activities
        opportunity_id = opportunity.id
    
    # === PROCESS TASK (optional) ===
    next_task_title = safe_str(row_data.get('next_task_title'))
    if next_task_title:
        task = Task(
            id=generate_id(),
            opportunity_id=opportunity.id,
            title=next_task_title,
            status='open',
            created_at=get_iso_timestamp(),
            updated_at=get_iso_timestamp()
        )
        
        # Due date
        due_date_str = row_data.get('next_task_due_date')
        if due_date_str:
            due_date, err = parse_date(due_date_str)
            if due_date:
                task.due_date = due_date
            else:
                warnings.append(f"Invalid due_date: {err}")
        
        # HOTFIX 6.1: Template - try next_task_template first, fallback to task_template
        task_template_name = safe_str(row_data.get('next_task_template')) or safe_str(row_data.get('task_template'))
        if task_template_name:
            template = config_cache['task_templates_by_name'].get(task_template_name.lower())
            if template:
                task.task_template_id = template.id
                # If no due_date but template has default_due_days, calculate
                if not task.due_date and template.default_due_days:
                    from datetime import timedelta
                    task.due_date = (datetime.now() + timedelta(days=template.default_due_days)).strftime('%Y-%m-%d')
        
        # Assigned to
        if owner_email and owner_email in config_cache['users_by_email']:
            task.assigned_to_user_id = config_cache['users_by_email'][owner_email.lower()].id
        
        if not dry_run:
            db.add(task)
    
    # === PROCESS ACTIVITY (optional) ===
    last_activity_text = safe_str(row_data.get('last_activity_text'))
    if last_activity_text:
        # Truncate if too long
        if len(last_activity_text) > 500:
            last_activity_text = last_activity_text[:497] + '...'
            warnings.append("Activity text truncated to 500 chars")
        
        # HOTFIX 6.1: Use last_activity_date if available
        occurred_at = get_iso_timestamp()  # default: now
        activity_date_str = row_data.get('last_activity_date')
        if activity_date_str:
            activity_date, err = parse_date(activity_date_str)
            if activity_date:
                occurred_at = activity_date + 'T12:00:00Z'  # Add time component
            else:
                warnings.append(f"Invalid last_activity_date: {err}, using current time")
        
        activity = Activity(
            id=generate_id(),
            opportunity_id=opportunity.id,
            type='note',
            occurred_at=occurred_at,
            summary=last_activity_text,
            created_by_user_id=current_user.id,
            created_at=get_iso_timestamp()
        )
        
        if not dry_run:
            db.add(activity)
    
    return ImportRowResult(
        row=row_num,
        status="imported",
        warnings=warnings,
        errors=[],
        account_name=account_name,
        opportunity_id=opportunity_id
    )


@router.post("/excel", response_model=ImportResponse)
async def import_excel(
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin", "sales"))
):
    """
    Import opportunities from Excel file
    
    **PASO 6 + HOTFIX 6.1** - Robust Excel importer
    
    **Permissions:** admin, sales
    
    File format: IMPORT_NORMALIZADO_CRM.xlsx, sheet "DATA"
    
    Columns expected:
    - account_name (required)
    - region, region_other, customer_type, customer_type_other
    - lead_source, lead_source_detail
    - owner_email
    - contact_first_name, contact_last_name, contact_role
    - contact_email, contact_phone
    - opportunity_name
    - stage (required)
    - expected_value_eur (required)
    - weighted_value_eur (optional)
    - forecast_close_month
    - close_outcome, close_date, won_value_eur, lost_reason
    - next_task_title, next_task_due_date, next_task_template
    - last_activity_text, last_activity_date
    
    Rules:
    - Account: upsert by name (case-insensitive)
    - Contact: create if not exists
    - Opportunity: always create (no upsert)
    - Task, Activity: create if data present
    - NO automatic cfg_* creation
    - Values not in catalog → save to *_other_text + warning
    - Empty rows: automatically skipped
    
    HOTFIX 6.1 changes:
    - Empty rows are now skipped (not counted as failed)
    - Column mapping: next_task_template (with fallback to task_template)
    - Column mapping: last_activity_date for activity occurred_at
    - region_other and customer_type_other properly used
    - weighted_value_eur support added
    
    dry_run=true: validate only, no DB writes
    """
    import_id = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Save uploaded file
    file_path = UPLOAD_DIR / f"import_{import_id}_{file.filename}"
    with file_path.open('wb') as f:
        shutil.copyfileobj(file.file, f)
    
    logger.info(f"Import {import_id} started by {current_user.email}, dry_run={dry_run}")
    
    # Load Excel
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error reading Excel file: {str(e)}"
        )
    
    if 'DATA' not in wb.sheetnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sheet 'DATA' not found in Excel file"
        )
    
    ws = wb['DATA']
    
    # Read header row (row 1)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col_idx
    
    # Preload config data (cache)
    config_cache = {
        'stages_by_key': {},
        'lead_sources_by_name': {},
        'regions_by_name': {},
        'customer_types_by_name': {},
        'contact_roles_by_name': {},
        'task_templates_by_name': {},
        'users_by_email': {}
    }
    
    stages = db.query(CfgStage).filter(CfgStage.is_active == 1).all()
    for s in stages:
        config_cache['stages_by_key'][s.key.lower()] = s
    
    lead_sources = db.query(CfgLeadSource).filter(CfgLeadSource.is_active == 1).all()
    for ls in lead_sources:
        config_cache['lead_sources_by_name'][ls.name.lower()] = ls
    
    regions = db.query(CfgRegion).filter(CfgRegion.is_active == 1).all()
    for r in regions:
        config_cache['regions_by_name'][r.name.lower()] = r
    
    customer_types = db.query(CfgCustomerType).filter(CfgCustomerType.is_active == 1).all()
    for ct in customer_types:
        config_cache['customer_types_by_name'][ct.name.lower()] = ct
    
    contact_roles = db.query(CfgContactRole).filter(CfgContactRole.is_active == 1).all()
    for cr in contact_roles:
        config_cache['contact_roles_by_name'][cr.name.lower()] = cr
    
    task_templates = db.query(CfgTaskTemplate).filter(CfgTaskTemplate.is_active == 1).all()
    for tt in task_templates:
        config_cache['task_templates_by_name'][tt.name.lower()] = tt
    
    users = db.query(User).filter(User.is_active == 1).all()
    for u in users:
        config_cache['users_by_email'][u.email.lower()] = u
    
    # Process rows
    results = []
    total_rows = 0
    imported_rows = 0
    failed_rows = 0
    skipped_rows = 0
    warnings_count = 0
    errors_count = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Build row_data dict
        row_data = {}
        for col_name, col_idx in headers.items():
            cell_value = row[col_idx - 1].value
            row_data[col_name] = cell_value
        
        # === SKIP EMPTY ROWS (HOTFIX 6.1) ===
        # Check if row is completely empty (critical fields)
        account_name_val = safe_str(row_data.get('account_name'))
        stage_val = safe_str(row_data.get('stage'))
        expected_value_val = row_data.get('expected_value_eur')
        
        # If all critical fields are empty, skip this row
        if not account_name_val and not stage_val and not expected_value_val:
            # Check if ALL fields are empty (truly empty row)
            all_empty = all(
                v is None or (isinstance(v, str) and not v.strip())
                for v in row_data.values()
            )
            
            if all_empty:
                # Skip completely empty row, don't count as total
                continue
            else:
                # Some fields have data but critical ones don't - count as skipped
                skipped_rows += 1
                total_rows += 1
                results.append(ImportRowResult(
                    row=row_idx,
                    status="skipped",
                    warnings=["Row skipped: critical fields (account_name, stage, expected_value_eur) are empty"],
                    errors=[],
                    account_name=None,
                    opportunity_id=None
                ))
                continue
        
        total_rows += 1
        
        # Process row
        result = process_excel_row(
            row_data=row_data,
            row_num=row_idx,
            db=db,
            current_user=current_user,
            dry_run=dry_run,
            config_cache=config_cache
        )
        
        results.append(result)
        
        if result.status == "imported":
            imported_rows += 1
        elif result.status == "failed":
            failed_rows += 1
        elif result.status == "skipped":
            skipped_rows += 1
        
        warnings_count += len(result.warnings)
        errors_count += len(result.errors)
        
        # Limit results to 200 in response
        if len(results) >= 200:
            break
    
    # Commit if not dry_run
    if not dry_run:
        try:
            db.commit()
            logger.info(f"Import {import_id} committed: {imported_rows} imported, {failed_rows} failed")
        except Exception as e:
            db.rollback()
            logger.error(f"Import {import_id} rollback: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Database error during import: {str(e)}"
            )
    
    # Create audit log
    if not dry_run:
        create_audit_log(
            db=db,
            entity="imports",
            entity_id=import_id,
            action="import",
            user_id=current_user.id,
            after_data={
                "total_rows": total_rows,
                "imported_rows": imported_rows,
                "failed_rows": failed_rows,
                "file": file.filename
            }
        )
        db.commit()
    
    # Save report JSON
    report_data = {
        "import_id": import_id,
        "dry_run": dry_run,
        "total_rows": total_rows,
        "imported_rows": imported_rows,
        "failed_rows": failed_rows,
        "skipped_rows": skipped_rows,
        "warnings_count": warnings_count,
        "errors_count": errors_count,
        "items": [r.dict() for r in results],
        "timestamp": get_iso_timestamp()
    }
    
    report_path = UPLOAD_DIR / f"import_report_{import_id}.json"
    with report_path.open('w') as f:
        json.dump(report_data, f, indent=2)
    
    return ImportResponse(
        import_id=import_id,
        dry_run=dry_run,
        total_rows=total_rows,
        imported_rows=imported_rows,
        failed_rows=failed_rows,
        skipped_rows=skipped_rows,
        warnings_count=warnings_count,
        errors_count=errors_count,
        items=results[:200],  # Limit to 200 in response
        file_saved=str(file_path),  # HOTFIX 6.1: Always save file path (even in dry_run for audit)
        report_saved=str(report_path)
    )


@router.get("/report/{import_id}", response_model=ImportReportResponse)
def get_import_report(
    import_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_from_cookie)
):
    """
    Get saved import report
    
    **Permissions:** All authenticated users
    """
    report_path = UPLOAD_DIR / f"import_report_{import_id}.json"
    
    if not report_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Import report {import_id} not found"
        )
    
    with report_path.open('r') as f:
        report_data = json.load(f)
    
    return ImportReportResponse(**report_data)
